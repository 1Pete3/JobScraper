import datetime
import os
import pickle
import time
import re

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

pages = 5
hcaptcha = True
siteURL = "https://www.higheredjobs.com/admin/search.cfm?JobCat=242&CatName=Computer%20and%20Information%20Technology"
cookieFile = "./cookies.pkl"
jobPage = []
salaries = []

def getPageURLS():
    arrURLS = [siteURL]
    for i in range(2, pages + 1):
        arrURLS.append(siteURL + "&Page=" + str(i))
    return arrURLS


def deleteCookies():
    if os.path.exists(cookieFile):
        os.remove(cookieFile)
        print(f"Deleted cookie file: {cookieFile}")
    else:
        print("Cookie file does not exist.")


def findCookies(driver, cookieFile):
    try:
        # Load cookies from the pickle file
        with open(cookieFile, "rb") as file:
            cookies = pickle.load(file)
            for cookie in cookies:
                # Add each cookie to the driver
                driver.add_cookie(cookie)

        # Refresh the page to use the loaded cookies

        print("Cookies loaded and page refreshed. 🍪")
        return True

    except FileNotFoundError:
        print("Without my cookies I'm just a monster 😔.")
        return False
    except Exception as e:
        print(f"Error loading cookies: {e}")
        return False


def saveCookies(driver):
    try:

        WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="js-results"]'))
        )
        cookies = driver.get_cookies()
        with open("cookies.pkl", "wb") as f:
            pickle.dump(cookies, f)
    except TimeoutException:
        print("Solve the hcaptcha faster")
        exit()
    except NoSuchElementException:
        print("Element not found")
        exit()


def checkCookies(driver):
    if findCookies(driver, cookieFile) == False:
        saveCookies(driver)

    else:
        try:
            element = driver.find_element(By.XPATH, '//*[@id="main-iframe"]')
            hcaptcha_url = element.get_attribute("src")
            print("hCaptcha URL:", hcaptcha_url)
            driver.delete_all_cookies()
            deleteCookies()
            driver.refresh()
            time.sleep(20)
            saveCookies(driver)
        except NoSuchElementException:
            print("No need for hCaptcha")
            hcaptcha = False


def scrape():
    wbName = "job_data_with_links.xlsx"
    pagesToScrape = getPageURLS()
    options = Options()
    options.add_argument("--log-level=3")
    options.add_argument("--disable-infobars")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    )

    driver = webdriver.Chrome(options=options)
    wb = Workbook()
    ws = wb.active
    job_data = []
    for i in range(len(pagesToScrape)):
        try:
            driver.get(pagesToScrape[i])
            checkCookies(driver)

            jobs = WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.ID, "page" + str(i + 1)))
            )
            time.sleep(3)
            data = jobs.get_attribute("outerHTML")
            soup = BeautifulSoup(data, "html.parser")
            job_entries = soup.find_all("div", class_="row record")

            for job_div in job_entries:
                # Extract job link and title
                job_link = job_div.find("a")["href"]
                job_title = job_div.find("a").get_text(strip=True)

                # Extract university name and location
                university_info = (
                    job_div.find("div", class_="col-sm-7")
                    .get_text(separator="\n", strip=True)
                    .split("\n")
                )
                university_name = university_info[1]  # university name
                location = university_info[2]  # Location

                # Extract job type and posted date
                job_type_info = (
                    job_div.find("div", class_="col-sm-5 text-sm-right")
                    .get_text(separator="\n", strip=True)
                    .split("\n")
                )
                job_type = job_type_info[0]  # Job type
                posted_date = job_type_info[1].replace("Posted ", "")  # Posted date

                # Check for salary information
                salary_span = job_div.find("span", class_="job-salary")
                salary = (
                    salary_span.get_text(strip=True) if salary_span else "Not listed"
                )
                salaries.append(salary)
               

                job_data.append(
                    {

                        "Job Link": job_link,
                        "Title": job_title,
                        "University": university_name,
                        "Location": location,
                        "Job Type": job_type,
                        "Posted Date": posted_date,
                        "Salary": salary,
                    }
                )

                df = pd.DataFrame(job_data)
                df.rename(
                    columns={
                        "Title": "Title",
                        "University": "University",
                        "Location": "Location",
                        "Job Type": "Job Type",
                        "Posted Date": "Posted Date",
                        "Salary": "Salary",
                    },
                    inplace=True,
                )

                ws.append(
                    [
                        "Job Title",
                        "University",
                        "Location",
                        "Type",
                        "Date Posted",
                        "Salary",
                    ]
                )

            # Write data and create hyperlinks for job titles
            for index, row in df.iterrows():
                cell = ws.cell(row=index + 2, column=1, value=row["Title"])
                cell.hyperlink = Hyperlink(
                    ref=cell.coordinate,
                    target="https://www.higheredjobs.com/admin/"
                    + row["Job Link"]
                    + row["Title"],
                    tooltip="Click to view job",
                )

                # Set font color to blue and underline the hyperlink
                cell.font = Font(color="0000FF", underline="single")
                ws.cell(row=index + 2, column=2, value=row["University"])
                ws.cell(row=index + 2, column=3, value=row["Location"])
                ws.cell(row=index + 2, column=4, value=row["Job Type"])
                ws.cell(row=index + 2, column=5, value=row["Posted Date"])
                ws.cell(row=index + 2, column=6, value=row["Salary"])

            # Formatting column names
            for column in range(1, len(ws[1]) + 1):
                col_letter = get_column_letter(column)
                max_length = max(
                    len(str(cell.value)) for cell in ws[col_letter] if cell.value
                )
                adjusted_width = max_length + 2  # Add some extra space for padding
                ws.column_dimensions[col_letter].width = adjusted_width
        finally:
            print("done")
    wb.save(wbName)
    print(salaries)
    #for i in range(len(job_data)):
     #   print(job_data[i]["Job Link"])
    # cprint(job_data)


scrape()
